"""Vues pour la gestion des congés."""

import json
from datetime import datetime, timedelta

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import Count, Q
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    ListView,
    UpdateView,
)

from accounts.models import User

from .forms import LeaveCalendarForm, LeaveRequestForm
from .models import FrenchHoliday, LeaveCalendar, LeaveRequest


class AccueilRequiredMixin(UserPassesTestMixin):
    """Mixin qui vérifie si l'utilisateur a le rôle 'Accueil'."""

    def test_func(self):
        return self.request.user.user_roles.filter(role__name="Accueil", is_active=True).exists()


class LeaveCalendarListView(LoginRequiredMixin, ListView):
    """Liste des calendriers de congés."""

    model = LeaveCalendar
    template_name = "leave/calendar_list.html"
    context_object_name = "calendars"

    def get_queryset(self):
        """Filtre les calendriers actifs."""
        today = timezone.now().date()
        # Vérifier si l'utilisateur est du rôle Accueil
        is_accueil = self.request.user.user_roles.filter(
            role__name="Accueil", is_active=True
        ).exists()
        
        if is_accueil:
            # Les Accueil voient tous les calendriers actifs
            return LeaveCalendar.objects.filter(
                is_active=True
            ).order_by("-start_date")
        else:
            # Les autres voient seulement les calendriers actuellement visibles
            return LeaveCalendar.objects.filter(
                Q(is_active=True)
                & Q(start_date__lte=today)
                & Q(visibility_end_date__gte=today)
            ).order_by("-start_date")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()

        # Vérifier si l'utilisateur est du rôle Accueil
        context["is_accueil"] = self.request.user.user_roles.filter(role__name="Accueil", is_active=True).exists()

        # Liste des anciens calendriers (pour Accueil uniquement)
        if context["is_accueil"]:
            context["archived_calendars"] = LeaveCalendar.objects.filter(
                visibility_end_date__lt=today
            ).order_by("-start_date")[:10]

        return context


class LeaveCalendarDetailView(LoginRequiredMixin, DetailView):
    """Détail d'un calendrier avec vue calendrier et tableau."""

    model = LeaveCalendar
    template_name = "leave/calendar_detail.html"
    context_object_name = "calendar"
    pk_url_kwarg = "pk"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        calendar = self.get_object()

        # Toutes les demandes pour ce calendrier
        requests = LeaveRequest.objects.filter(calendar=calendar).select_related("user")

        # Regroupement par date
        requests_by_date = {}
        for req in requests:
            date_str = req.date.strftime("%Y-%m-%d")
            if date_str not in requests_by_date:
                requests_by_date[date_str] = []
            requests_by_date[date_str].append({
                "id": str(req.id),
                "user": f"{req.user.first_name} {req.user.last_name}",
                "period": req.period,
                "notes": req.notes,
            })

        context["requests_by_date"] = json.dumps(requests_by_date)
        context["is_accueil"] = self.request.user.user_roles.filter(role__name="Accueil", is_active=True).exists()

        # Données pour le tableau récapitulatif
        context["all_requests"] = requests.order_by("date", "user__last_name")

        # Liste des utilisateurs ayant fait des demandes
        users_with_requests = (
            User.objects.filter(leave_requests__calendar=calendar)
            .distinct()
            .prefetch_related("user_roles__role")
        )
        context["users_with_requests"] = users_with_requests

        # Jours fériés
        holidays = FrenchHoliday.objects.filter(
            date__gte=calendar.start_date,
            date__lte=calendar.end_date,
            is_active=True,
        )
        context["holidays"] = {h.date.strftime("%Y-%m-%d"): h.name for h in holidays}

        return context


class LeaveCalendarCreateView(LoginRequiredMixin, AccueilRequiredMixin, CreateView):
    """Création d'un calendrier de congés (rôle Accueil uniquement)."""

    model = LeaveCalendar
    form_class = LeaveCalendarForm
    template_name = "leave/calendar_form.html"
    success_url = reverse_lazy("leave:calendar_list")

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)

        messages.success(
            self.request,
            f"Le calendrier '{form.instance.name}' a été créé avec succès. "
            "Les agents seront notifiés.",
        )

        # TODO: Envoyer notification à tous les agents
        return response


class LeaveCalendarUpdateView(LoginRequiredMixin, AccueilRequiredMixin, UpdateView):
    """Modification d'un calendrier (rôle Accueil uniquement)."""

    model = LeaveCalendar
    form_class = LeaveCalendarForm
    template_name = "leave/calendar_form.html"
    pk_url_kwarg = "pk"

    def get_success_url(self):
        return self.object.get_absolute_url()

    def form_valid(self, form):
        messages.success(self.request, "Le calendrier a été mis à jour.")
        return super().form_valid(form)


class LeaveCalendarDeleteView(LoginRequiredMixin, AccueilRequiredMixin, DeleteView):
    """Suppression d'un calendrier (rôle Accueil uniquement)."""

    model = LeaveCalendar
    template_name = "leave/calendar_delete_confirm.html"
    success_url = reverse_lazy("leave:calendar_list")
    pk_url_kwarg = "pk"

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Le calendrier a été supprimé.")
        return super().delete(request, *args, **kwargs)


class LeaveRequestCreateView(LoginRequiredMixin, CreateView):
    """Pose de congés."""

    model = LeaveRequest
    form_class = LeaveRequestForm
    template_name = "leave/leave_request_form.html"

    def dispatch(self, request, *args, **kwargs):
        self.calendar = get_object_or_404(
            LeaveCalendar, pk=self.kwargs.get("calendar_pk")
        )
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Vérifier si l'utilisateur est du rôle Accueil
        context["is_accueil"] = self.request.user.user_roles.filter(
            role__name="Accueil"
        ).exists()

        return context

    def form_valid(self, form):
        selected_dates = form.cleaned_data["selected_dates"]
        period = form.cleaned_data["period"]
        notes = form.cleaned_data["notes"]

        created_count = 0
        errors = []

        for date_str in selected_dates:
            try:
                date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()

                # Vérifier si la date est dans la période du calendrier
                if not (self.calendar.start_date <= date_obj <= self.calendar.end_date):
                    errors.append(f"{date_str}: hors de la période du calendrier")
                    continue

                # Vérifier si une demande existe déjà
                if LeaveRequest.objects.filter(
                    calendar=self.calendar,
                    user=self.request.user,
                    date=date_obj,
                    period=period,
                ).exists():
                    errors.append(f"{date_str}: demande déjà existante")
                    continue

                LeaveRequest.objects.create(
                    calendar=self.calendar,
                    user=self.request.user,
                    date=date_obj,
                    period=period,
                    notes=notes,
                )
                created_count += 1

            except ValueError:
                errors.append(f"{date_str}: format de date invalide")

        if created_count > 0:
            messages.success(
                self.request,
                f"{created_count} jour(s) de congés posé(s) avec succès.",
            )

        if errors:
            for error in errors:
                messages.warning(self.request, error)

        return redirect("leave:calendar_detail", pk=self.calendar.pk)

    def form_invalid(self, form):
        messages.error(
            self.request, "Veuillez corriger les erreurs dans le formulaire."
        )
        return super().form_invalid(form)


class LeaveRequestUpdateView(LoginRequiredMixin, UpdateView):
    """Modification d'une demande de congés (propriétaire uniquement)."""

    model = LeaveRequest
    form_class = LeaveRequestForm
    template_name = "leave/leave_request_update.html"
    pk_url_kwarg = "pk"

    def get_queryset(self):
        """Seul le propriétaire peut modifier sa demande."""
        return LeaveRequest.objects.filter(user=self.request.user)

    def form_valid(self, form):
        messages.success(self.request, "Votre demande de congés a été modifiée.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse("leave:user_requests")


class LeaveRequestDeleteView(LoginRequiredMixin, DeleteView):
    """Suppression d'une demande de congés (propriétaire uniquement)."""

    model = LeaveRequest
    template_name = "leave/leave_request_delete_confirm.html"
    pk_url_kwarg = "pk"

    def get_queryset(self):
        """Seul le propriétaire peut supprimer sa demande."""
        return LeaveRequest.objects.filter(user=self.request.user)

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Votre demande de congés a été supprimée.")
        return super().delete(request, *args, **kwargs)

    def get_success_url(self):
        return self.object.calendar.get_absolute_url()


class LeaveCalendarEventsAPI(LoginRequiredMixin, View):
    """API pour récupérer les événements du calendrier (JSON pour FullCalendar)."""

    def get(self, request, pk, *args, **kwargs):
        calendar = get_object_or_404(LeaveCalendar, pk=pk)

        # Récupérer toutes les demandes
        requests = LeaveRequest.objects.filter(calendar=calendar).select_related("user")

        events = []
        for req in requests:
            # Couleur fixe pour tous les événements
            color = "#3788d8"

            # Titre avec le nom de l'utilisateur
            title = f"{req.user.first_name} {req.user.last_name}"
            if req.period != "full":
                title += f" ({req.get_period_display()})"

            events.append(
                {
                    "id": str(req.id),
                    "title": title,
                    "start": req.date.isoformat(),
                    "color": color,
                    "textColor": "#ffffff",
                    "extendedProps": {
                        "period": req.period,
                        "user": f"{req.user.first_name} {req.user.last_name}",
                        "notes": req.notes,
                    },
                }
            )

        # Ajouter les jours fériés
        holidays = FrenchHoliday.objects.filter(
            date__gte=calendar.start_date,
            date__lte=calendar.end_date,
            is_active=True,
        )

        for holiday in holidays:
            events.append(
                {
                    "id": f"holiday_{holiday.id}",
                    "title": holiday.name,
                    "start": holiday.date.isoformat(),
                    "color": "#dc3545",
                    "textColor": "#ffffff",
                    "display": "background",
                    "extendedProps": {"is_holiday": True},
                }
            )

        return JsonResponse(events, safe=False)


class LeaveExportExcelView(LoginRequiredMixin, AccueilRequiredMixin, View):
    """Export Excel d'un calendrier de congés."""

    def get(self, request, pk, *args, **kwargs):
        calendar = get_object_or_404(LeaveCalendar, pk=pk)

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            messages.error(
                request,
                "L'export Excel nécessite la bibliothèque openpyxl.",
            )
            return redirect("leave:calendar_detail", pk=pk)

        wb = Workbook()
        ws = wb.active
        ws.title = "Récapitulatif"

        # En-têtes
        headers = ["Nom"]

        # Générer les dates du calendrier (semaines uniquement)
        current_date = calendar.start_date
        dates = []
        while current_date <= calendar.end_date:
            if current_date.weekday() < 5:  # Lundi = 0, Vendredi = 4
                dates.append(current_date)
                headers.append(current_date.strftime("%d/%m"))
            current_date += timedelta(days=1)

        # Écrire les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Données
        users = (
            User.objects.filter(leave_requests__calendar=calendar)
            .distinct()
        )

        row = 2
        for user in users:
            # Nom
            ws.cell(row=row, column=1, value=f"{user.first_name} {user.last_name}")

            # Demandes par date
            user_requests = LeaveRequest.objects.filter(
                calendar=calendar, user=user
            ).values_list("date", "period")

            requests_dict = {date: period for date, period in user_requests}

            for col_idx, date in enumerate(dates, 2):
                if date in requests_dict:
                    cell = ws.cell(row=row, column=col_idx, value="X")
                    cell.fill = PatternFill(
                        start_color="FFA500", end_color="FFA500", fill_type="solid"
                    )
                    cell.alignment = Alignment(horizontal="center")

            row += 1

        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15

        # Réponse HTTP avec le fichier Excel
        from django.http import HttpResponse

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="conges_{calendar.name.replace(" ", "_")}.xlsx"'
        )
        wb.save(response)

        return response


class ArchivedCalendarsView(LoginRequiredMixin, AccueilRequiredMixin, ListView):
    """Liste des calendriers archivés (rôle Accueil uniquement)."""

    model = LeaveCalendar
    template_name = "leave/calendar_archives.html"
    context_object_name = "calendars"

    def get_queryset(self):
        """Calendriers dont la période de visibilité est terminée."""
        today = timezone.now().date()
        return LeaveCalendar.objects.filter(visibility_end_date__lt=today).order_by(
            "-start_date"
        )


class UserLeaveRequestsView(LoginRequiredMixin, ListView):
    """Vue des congés de l'utilisateur connecté avec interface en cartes."""

    model = LeaveRequest
    template_name = "leave/user_requests.html"
    context_object_name = "requests"

    def get_queryset(self):
        """Retourne uniquement les demandes de l'utilisateur connecté."""
        return LeaveRequest.objects.filter(
            user=self.request.user
        ).select_related("calendar").order_by("-date")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Grouper par calendrier pour l'affichage
        requests_by_calendar = {}
        for req in self.get_queryset():
            cal_id = str(req.calendar.id)
            if cal_id not in requests_by_calendar:
                requests_by_calendar[cal_id] = {
                    "calendar": req.calendar,
                    "requests": [],
                }
            requests_by_calendar[cal_id]["requests"].append(req)
        context["requests_by_calendar"] = requests_by_calendar
        return context


class BulkDeleteLeaveRequestsView(LoginRequiredMixin, View):
    """Vue pour suppression multiple des demandes de congés."""

    def post(self, request, *args, **kwargs):
        """Supprime les demandes sélectionnées."""
        request_ids = request.POST.getlist("request_ids[]")

        if not request_ids:
            messages.warning(request, "Aucune demande sélectionnée.")
            return redirect("leave:user_requests")

        # Vérifier que toutes les demandes appartiennent à l'utilisateur
        requests_to_delete = LeaveRequest.objects.filter(
            id__in=request_ids,
            user=request.user
        )

        deleted_count = requests_to_delete.count()

        if deleted_count == 0:
            messages.error(request, "Aucune demande trouvée ou vous n'avez pas les droits.")
            return redirect("leave:user_requests")

        # Suppression
        requests_to_delete.delete()

        messages.success(
            request,
            f"{deleted_count} congé{'s' if deleted_count > 1 else ''} supprimé{'s' if deleted_count > 1 else ''} avec succès."
        )

        return redirect("leave:user_requests")
